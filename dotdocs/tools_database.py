from __future__ import annotations

from contextlib import closing
from datetime import date, datetime, timezone
import json
from pathlib import Path
import re
import sqlite3

from openpyxl import Workbook, load_workbook


TOOLS_SCHEMA = """
CREATE TABLE IF NOT EXISTS tools (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    display_tool_id TEXT NOT NULL,
    normalized_tool_id TEXT NOT NULL UNIQUE,
    description TEXT NOT NULL,
    category TEXT NOT NULL DEFAULT '',
    manufacturer TEXT NOT NULL DEFAULT '',
    model TEXT NOT NULL DEFAULT '',
    serial_number TEXT NOT NULL DEFAULT '',
    normalized_serial_number TEXT NOT NULL DEFAULT '',
    location TEXT NOT NULL DEFAULT '',
    custodian TEXT NOT NULL DEFAULT '',
    calibration_required INTEGER NOT NULL DEFAULT 1 CHECK(calibration_required IN (0, 1)),
    calibration_interval_months INTEGER,
    active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0, 1)),
    notes TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);
CREATE UNIQUE INDEX IF NOT EXISTS idx_tools_serial_unique
    ON tools(normalized_serial_number) WHERE normalized_serial_number <> '';
CREATE INDEX IF NOT EXISTS idx_tools_active ON tools(active);

CREATE TABLE IF NOT EXISTS tool_certifications (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tool_id INTEGER NOT NULL REFERENCES tools(id) ON DELETE RESTRICT,
    certificate_type TEXT NOT NULL,
    certificate_number TEXT NOT NULL DEFAULT '',
    performed_date TEXT,
    due_date TEXT,
    provider TEXT NOT NULL DEFAULT '',
    result TEXT NOT NULL DEFAULT 'unknown' CHECK(result IN ('pass', 'fail', 'limited', 'unknown')),
    document_path TEXT NOT NULL DEFAULT '',
    document_sha256 TEXT NOT NULL DEFAULT '',
    source_review_id TEXT NOT NULL DEFAULT '',
    notes TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_tool_certifications_tool ON tool_certifications(tool_id, id DESC);
CREATE UNIQUE INDEX IF NOT EXISTS idx_tool_certifications_document_hash
    ON tool_certifications(document_sha256) WHERE document_sha256 <> '';

CREATE TABLE IF NOT EXISTS tool_database_audit (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp_utc TEXT NOT NULL,
    event TEXT NOT NULL,
    tool_id INTEGER,
    details_json TEXT NOT NULL
);
"""

TOOL_TEXT_FIELDS = (
    "description",
    "category",
    "manufacturer",
    "model",
    "serial_number",
    "location",
    "custodian",
    "notes",
)
CERTIFICATE_RESULTS = {"pass", "fail", "limited", "unknown"}
CERTIFICATE_TYPES = {"calibration": "Calibration", "certification": "Certification"}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _text(value: object, *, maximum: int = 500) -> str:
    text = str(value or "").strip()
    if len(text) > maximum:
        raise ValueError(f"Value exceeds the {maximum}-character limit.")
    return text


def normalize_tool_identifier(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def _connection(database_path: str | Path) -> sqlite3.Connection:
    path = Path(database_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(path, timeout=30)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    connection.executescript(TOOLS_SCHEMA)
    return connection


def ensure_tools_schema(database_path: str | Path) -> None:
    with closing(_connection(database_path)) as connection:
        connection.commit()


def _tool_record(row: sqlite3.Row) -> dict:
    record = dict(row)
    record["calibration_required"] = bool(record["calibration_required"])
    record["active"] = bool(record["active"])
    return record


def _validated_tool(values: dict, current: dict | None = None) -> dict:
    current = current or {}
    display_tool_id = _text(values.get("tool_id", values.get("display_tool_id", current.get("display_tool_id"))), maximum=60)
    normalized_tool_id = normalize_tool_identifier(display_tool_id)
    if not normalized_tool_id:
        raise ValueError("Tool ID must contain at least one letter or number.")
    cleaned = {field: _text(values.get(field, current.get(field)), maximum=1000 if field == "notes" else 120) for field in TOOL_TEXT_FIELDS}
    if not cleaned["description"]:
        raise ValueError("Tool description is required.")
    normalized_serial = normalize_tool_identifier(cleaned["serial_number"])
    required = bool(values.get("calibration_required", current.get("calibration_required", True)))
    active = bool(values.get("active", current.get("active", True)))
    interval_value = values.get("calibration_interval_months", current.get("calibration_interval_months"))
    if interval_value in (None, ""):
        interval = None
    else:
        try:
            interval = int(interval_value)
        except (TypeError, ValueError) as error:
            raise ValueError("Calibration interval must be a positive whole number of months.") from error
        if interval < 1:
            raise ValueError("Calibration interval must be a positive whole number of months.")
    return {
        "display_tool_id": display_tool_id,
        "normalized_tool_id": normalized_tool_id,
        **cleaned,
        "normalized_serial_number": normalized_serial,
        "calibration_required": required,
        "calibration_interval_months": interval,
        "active": active,
    }


def _audit(connection: sqlite3.Connection, event: str, tool_id: int | None, details: dict) -> None:
    connection.execute(
        "INSERT INTO tool_database_audit(timestamp_utc, event, tool_id, details_json) VALUES (?, ?, ?, ?)",
        (_now(), event, tool_id, json.dumps(details, sort_keys=True)),
    )


def _raise_integrity(error: sqlite3.IntegrityError) -> None:
    message = str(error).lower()
    if "normalized_tool_id" in message:
        raise ValueError("That Tool ID is already assigned to another tool.") from error
    if "normalized_serial_number" in message:
        raise ValueError("That serial number is already assigned to another tool.") from error
    if "document_sha256" in message:
        raise ValueError("That certificate document is already linked to calibration history.") from error
    raise ValueError("The tool record conflicts with existing database data.") from error


def create_tool(database_path: str | Path, values: dict) -> dict:
    cleaned = _validated_tool(values)
    timestamp = _now()
    columns = tuple(cleaned)
    try:
        with closing(_connection(database_path)) as connection:
            connection.execute("BEGIN IMMEDIATE")
            cursor = connection.execute(
                f"INSERT INTO tools ({', '.join(columns)}, created_at, updated_at) VALUES ({', '.join('?' for _ in columns)}, ?, ?)",
                (*cleaned.values(), timestamp, timestamp),
            )
            tool_id = cursor.lastrowid
            _audit(connection, "tool_created", tool_id, cleaned)
            connection.commit()
    except sqlite3.IntegrityError as error:
        _raise_integrity(error)
    return get_tool(database_path, tool_id)


def get_tool(database_path: str | Path, tool_id: int) -> dict:
    with closing(_connection(database_path)) as connection:
        row = connection.execute("SELECT * FROM tools WHERE id = ?", (tool_id,)).fetchone()
    if row is None:
        raise ValueError("The selected tool no longer exists.")
    return _tool_record(row)


def get_tool_by_identifier(database_path: str | Path, tool_identifier: object) -> dict:
    normalized = normalize_tool_identifier(tool_identifier)
    if not normalized:
        raise ValueError("A Tool ID is required.")
    with closing(_connection(database_path)) as connection:
        row = connection.execute(
            "SELECT * FROM tools WHERE normalized_tool_id = ? AND active = 1",
            (normalized,),
        ).fetchone()
    if row is None:
        raise ValueError("No active tool has that Tool ID.")
    return _tool_record(row)


def list_tools(database_path: str | Path, search: object = "", *, include_inactive: bool = True) -> list[dict]:
    query = str(search or "").strip().casefold()
    with closing(_connection(database_path)) as connection:
        sql = "SELECT * FROM tools" if include_inactive else "SELECT * FROM tools WHERE active = 1"
        rows = connection.execute(sql).fetchall()
    records = [_tool_record(row) for row in rows]
    if query:
        fields = ("display_tool_id", "description", "category", "manufacturer", "model", "serial_number", "location", "custodian")
        records = [record for record in records if query in " ".join(str(record[field]).casefold() for field in fields)]
    return sorted(records, key=lambda record: record["display_tool_id"].casefold())


def match_tool_in_text(database_path: str | Path, text: str) -> dict:
    records = list_tools(database_path, include_inactive=False)
    normalized_text = str(text or "").upper()
    evidence = []
    for label, field in (
        (r"(?:SERIAL(?:\s+NO\.?|\s+NUMBER)?|S/N)", "normalized_serial_number"),
        (r"(?:TOOL\s+ID|ASSET\s+TAG|ASSET\s+ID)", "normalized_tool_id"),
    ):
        for match in re.finditer(label + r"\s*[:#-]?\s*([A-Z0-9][A-Z0-9._/-]{2,59})", normalized_text):
            value = normalize_tool_identifier(match.group(1))
            matches = [record for record in records if record[field] and record[field] == value]
            if matches:
                evidence.extend(record["id"] for record in matches)
    unique_ids = sorted(set(evidence))
    if len(unique_ids) == 1:
        tool = next(record for record in records if record["id"] == unique_ids[0])
        return {"status": "unique", "tool": tool}
    if len(unique_ids) > 1:
        return {"status": "conflict", "tool": None}
    return {"status": "unmatched", "tool": None}


def update_tool(database_path: str | Path, tool_id: int, values: dict) -> dict:
    current = get_tool(database_path, tool_id)
    cleaned = _validated_tool(values, current)
    columns = tuple(cleaned)
    try:
        with closing(_connection(database_path)) as connection:
            connection.execute("BEGIN IMMEDIATE")
            cursor = connection.execute(
                f"UPDATE tools SET {', '.join(f'{column} = ?' for column in columns)}, updated_at = ? WHERE id = ?",
                (*cleaned.values(), _now(), tool_id),
            )
            if cursor.rowcount != 1:
                raise ValueError("The selected tool no longer exists.")
            _audit(connection, "tool_updated", tool_id, cleaned)
            connection.commit()
    except sqlite3.IntegrityError as error:
        _raise_integrity(error)
    return get_tool(database_path, tool_id)


def _iso_date(value: object, label: str) -> str | None:
    text = _text(value, maximum=10)
    if not text:
        return None
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as error:
        raise ValueError(f"{label} must use YYYY-MM-DD.") from error


def _validated_certification(values: dict) -> dict:
    type_value = _text(values.get("certificate_type"), maximum=40).lower()
    if type_value not in CERTIFICATE_TYPES:
        raise ValueError("Certificate type must be Calibration or Certification.")
    performed = _iso_date(values.get("performed_date"), "Performed date")
    due = _iso_date(values.get("due_date"), "Due date")
    if performed and due and due < performed:
        raise ValueError("Due date cannot be before the performed date.")
    result = _text(values.get("result") or "unknown", maximum=20).lower()
    if result not in CERTIFICATE_RESULTS:
        raise ValueError("Calibration result must be pass, fail, limited, or unknown.")
    document_hash = _text(values.get("document_sha256"), maximum=64).lower()
    if document_hash and not re.fullmatch(r"[0-9a-f]{64}", document_hash):
        raise ValueError("Certificate document fingerprint must be a SHA-256 value.")
    return {
        "certificate_type": CERTIFICATE_TYPES[type_value],
        "certificate_number": _text(values.get("certificate_number"), maximum=120),
        "performed_date": performed,
        "due_date": due,
        "provider": _text(values.get("provider"), maximum=120),
        "result": result,
        "document_path": _text(values.get("document_path"), maximum=1000),
        "document_sha256": document_hash,
        "source_review_id": _text(values.get("source_review_id"), maximum=200),
        "notes": _text(values.get("notes"), maximum=1000),
    }


def add_tool_certification(database_path: str | Path, tool_id: int, values: dict) -> dict:
    cleaned = _validated_certification(values)
    columns = tuple(cleaned)
    try:
        with closing(_connection(database_path)) as connection:
            connection.execute("BEGIN IMMEDIATE")
            if connection.execute("SELECT 1 FROM tools WHERE id = ?", (tool_id,)).fetchone() is None:
                raise ValueError("The selected tool no longer exists.")
            cursor = connection.execute(
                f"INSERT INTO tool_certifications (tool_id, {', '.join(columns)}, created_at) VALUES (?, {', '.join('?' for _ in columns)}, ?)",
                (tool_id, *cleaned.values(), _now()),
            )
            certification_id = cursor.lastrowid
            _audit(connection, "tool_certification_added", tool_id, {"certification_id": certification_id, **cleaned})
            connection.commit()
    except sqlite3.IntegrityError as error:
        _raise_integrity(error)
    return get_tool_certification(database_path, certification_id)


def get_tool_certification(database_path: str | Path, certification_id: int) -> dict:
    with closing(_connection(database_path)) as connection:
        row = connection.execute("SELECT * FROM tool_certifications WHERE id = ?", (certification_id,)).fetchone()
    if row is None:
        raise ValueError("The selected calibration record no longer exists.")
    return dict(row)


def remove_tool_certification(database_path: str | Path, certification_id: int) -> None:
    with closing(_connection(database_path)) as connection:
        connection.execute("BEGIN IMMEDIATE")
        row = connection.execute(
            "SELECT tool_id FROM tool_certifications WHERE id = ?", (certification_id,)
        ).fetchone()
        if row is None:
            connection.rollback()
            return
        connection.execute("DELETE FROM tool_certifications WHERE id = ?", (certification_id,))
        _audit(connection, "tool_certification_rollback_removed", row["tool_id"], {"certification_id": certification_id})
        connection.commit()


def list_tool_certifications(database_path: str | Path, tool_id: int) -> list[dict]:
    with closing(_connection(database_path)) as connection:
        rows = connection.execute(
            "SELECT * FROM tool_certifications WHERE tool_id = ? ORDER BY id DESC",
            (tool_id,),
        ).fetchall()
    return [dict(row) for row in rows]


TOOL_WORKBOOK_HEADERS = {
    "Tool ID": "tool_id",
    "Description": "description",
    "Category": "category",
    "Manufacturer": "manufacturer",
    "Model": "model",
    "Serial Number": "serial_number",
    "Location": "location",
    "Custodian": "custodian",
    "Calibration Required": "calibration_required",
    "Interval Months": "calibration_interval_months",
    "Active": "active",
    "Notes": "notes",
}


def _workbook_boolean(value: object, label: str, *, default: bool) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return default
    choices = {"yes": True, "true": True, "1": True, "no": False, "false": False, "0": False}
    if text not in choices:
        raise ValueError(f"{label} must be Yes or No.")
    return choices[text]


def import_tools_workbook(workbook_path: str | Path, database_path: str | Path) -> dict[str, int]:
    """Atomically insert or update tools from a dedicated Tools worksheet."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    inserted = 0
    updated = 0
    try:
        sheet = workbook["Tools"] if "Tools" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        if "Tool ID" not in headers or "Description" not in headers:
            raise ValueError("Tools workbook must include Tool ID and Description columns.")
        rows = []
        seen_ids: set[str] = set()
        seen_serials: set[str] = set()
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            source = dict(zip(headers, values))
            if not any(value not in (None, "") for value in values):
                continue
            mapped = {field: source.get(header) for header, field in TOOL_WORKBOOK_HEADERS.items()}
            mapped["calibration_required"] = _workbook_boolean(
                mapped["calibration_required"], f"Calibration Required on row {row_number}", default=True,
            )
            mapped["active"] = _workbook_boolean(mapped["active"], f"Active on row {row_number}", default=True)
            cleaned = _validated_tool(mapped)
            if cleaned["normalized_tool_id"] in seen_ids:
                raise ValueError(f"Duplicate Tool ID in workbook row {row_number}.")
            serial = cleaned["normalized_serial_number"]
            if serial and serial in seen_serials:
                raise ValueError(f"Duplicate serial number in workbook row {row_number}.")
            seen_ids.add(cleaned["normalized_tool_id"])
            if serial:
                seen_serials.add(serial)
            rows.append(cleaned)

        with closing(_connection(database_path)) as connection:
            connection.execute("BEGIN IMMEDIATE")
            try:
                for cleaned in rows:
                    existing = connection.execute(
                        "SELECT id FROM tools WHERE normalized_tool_id = ?",
                        (cleaned["normalized_tool_id"],),
                    ).fetchone()
                    columns = tuple(cleaned)
                    if existing is None:
                        timestamp = _now()
                        cursor = connection.execute(
                            f"INSERT INTO tools ({', '.join(columns)}, created_at, updated_at) VALUES ({', '.join('?' for _ in columns)}, ?, ?)",
                            (*cleaned.values(), timestamp, timestamp),
                        )
                        tool_id = cursor.lastrowid
                        inserted += 1
                        event = "tool_imported"
                    else:
                        tool_id = existing["id"]
                        connection.execute(
                            f"UPDATE tools SET {', '.join(f'{column} = ?' for column in columns)}, updated_at = ? WHERE id = ?",
                            (*cleaned.values(), _now(), tool_id),
                        )
                        updated += 1
                        event = "tool_import_updated"
                    _audit(connection, event, tool_id, cleaned)
                connection.commit()
            except sqlite3.IntegrityError as error:
                connection.rollback()
                _raise_integrity(error)
    finally:
        workbook.close()
    return {"inserted": inserted, "updated": updated}


def export_tools_workbook(database_path: str | Path, output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    tools_sheet = workbook.active
    tools_sheet.title = "Tools"
    tools_sheet.append(list(TOOL_WORKBOOK_HEADERS))
    tools = list_tools(database_path)
    for tool in tools:
        tools_sheet.append([
            tool["display_tool_id"], tool["description"], tool["category"], tool["manufacturer"],
            tool["model"], tool["serial_number"], tool["location"], tool["custodian"],
            "Yes" if tool["calibration_required"] else "No", tool["calibration_interval_months"],
            "Yes" if tool["active"] else "No", tool["notes"],
        ])
    history = workbook.create_sheet("Certification History")
    history.append([
        "Tool ID", "Certificate Type", "Certificate Number", "Performed Date", "Due Date",
        "Provider", "Result", "Document Path", "Document SHA-256", "Notes",
    ])
    for tool in tools:
        for item in reversed(list_tool_certifications(database_path, tool["id"])):
            history.append([
                tool["display_tool_id"], item["certificate_type"], item["certificate_number"],
                item["performed_date"], item["due_date"], item["provider"], item["result"],
                item["document_path"], item["document_sha256"], item["notes"],
            ])
    workbook.save(output)
    workbook.close()
    return output
