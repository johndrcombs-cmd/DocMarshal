from __future__ import annotations

import sqlite3
import os
import tempfile
import json
import re
from collections import defaultdict
from contextlib import closing
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .normalization import normalize_plate, normalize_unit, normalize_vin

SCHEMA = """
CREATE TABLE units (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_row INTEGER NOT NULL,
    display_unit TEXT NOT NULL,
    normalized_unit TEXT NOT NULL,
    unit_type TEXT,
    year TEXT,
    make TEXT,
    model TEXT,
    vehicle_type TEXT,
    plate TEXT,
    vin TEXT,
    fuel_type TEXT,
    next_dot TEXT,
    dot_status TEXT,
    asset_owner TEXT,
    asset_source TEXT NOT NULL DEFAULT 'workbook'
);
CREATE INDEX idx_units_normalized_unit ON units(normalized_unit);
CREATE INDEX idx_units_plate ON units(plate);
CREATE INDEX idx_units_vin ON units(vin);
CREATE UNIQUE INDEX idx_manual_unit_unique ON units(normalized_unit)
    WHERE asset_source = 'manual' AND normalized_unit <> '';
CREATE UNIQUE INDEX idx_manual_plate_unique ON units(plate)
    WHERE asset_source = 'manual' AND plate <> '';
CREATE UNIQUE INDEX idx_manual_vin_unique ON units(vin)
    WHERE asset_source = 'manual' AND vin <> '';

CREATE TABLE import_issues (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_row INTEGER,
    issue_type TEXT NOT NULL,
    identifier_value TEXT,
    details TEXT
);
"""


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _create_database(database_path: Path) -> sqlite3.Connection:
    database_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(database_path)
    connection.executescript(SCHEMA)
    return connection


def import_fleet_workbook(
    workbook_path: str | Path,
    database_path: str | Path,
    *,
    manual_assets_path: str | Path | None = None,
) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    database_path = Path(database_path)
    database_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_file = tempfile.NamedTemporaryFile(
        dir=database_path.parent,
        prefix=database_path.name + ".",
        suffix=".tmp",
        delete=False,
    )
    temporary_database = Path(temporary_file.name)
    temporary_file.close()
    temporary_database.unlink()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [_text(cell.value) for cell in sheet[1]]

    imported = []
    manual_count = 0
    try:
        with closing(_create_database(temporary_database)) as connection:
            for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                record = dict(zip(headers, values))
                display_unit = _text(record.get("Unit #"))
                normalized_unit = normalize_unit(display_unit)
                plate = normalize_plate(record.get("Tag"))
                vin = normalize_vin(record.get("Vin"))
                connection.execute(
                    """
                    INSERT INTO units (
                        source_row, display_unit, normalized_unit, unit_type, year, make,
                        model, vehicle_type, plate, vin, fuel_type, next_dot, dot_status,
                        asset_owner, asset_source
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        source_row, display_unit, normalized_unit, _text(record.get("Unit Type")),
                        _text(record.get("Year")), _text(record.get("Make")), _text(record.get("Model")),
                        _text(record.get("Type")), plate, vin, _text(record.get("Fuel Type")),
                        _text(record.get("Next DOT")), _text(record.get("DOT Status")),
                        _text(record.get("Asset Owner")), "workbook",
                    ),
                )
                imported.append({"source_row": source_row, "unit": normalized_unit, "plate": plate, "vin": vin})

            if manual_assets_path is not None:
                from .assets import _validate_asset, load_manual_assets

                for saved_asset in load_manual_assets(manual_assets_path):
                    asset = _validate_asset(**saved_asset)
                    conflict = connection.execute(
                        """SELECT normalized_unit, plate, vin FROM units
                           WHERE normalized_unit = ? OR (? <> '' AND plate = ?) OR (? <> '' AND vin = ?)""",
                        (asset["unit"], asset["plate"], asset["plate"], asset["vin"], asset["vin"]),
                    ).fetchone()
                    if conflict:
                        raise ValueError(f"Manual asset {asset['unit']} conflicts with existing fleet data.")
                    connection.execute(
                        """INSERT INTO units (
                            source_row, display_unit, normalized_unit, unit_type, year, make,
                            model, vehicle_type, plate, vin, fuel_type, next_dot, dot_status,
                            asset_owner, asset_source
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (0, asset["unit"], asset["unit"], asset["unit_type"], asset["year"],
                         asset["make"], asset["model"], asset["vehicle_type"], asset["plate"],
                         asset["vin"], asset["fuel_type"], "", "", asset["asset_owner"], "manual"),
                    )
                    imported.append({"source_row": 0, "unit": asset["unit"], "plate": asset["plate"], "vin": asset["vin"]})
                    manual_count += 1

            duplicate_fields = {"plate": "AMBIGUOUS_PLATE", "vin": "AMBIGUOUS_VIN", "unit": "DUPLICATE_UNIT"}
            ambiguities: dict[str, dict[str, list[str]]] = {}
            for field, issue_type in duplicate_fields.items():
                grouped: dict[str, list[str]] = defaultdict(list)
                for record in imported:
                    identifier = record[field]
                    if identifier:
                        grouped[identifier].append(record["unit"])
                duplicate_values = {
                    identifier: sorted(set(units), key=lambda value: (len(value), value))
                    for identifier, units in grouped.items()
                    if len(units) > 1
                }
                ambiguities[field] = duplicate_values
                for identifier, units in duplicate_values.items():
                    connection.execute(
                        "INSERT INTO import_issues(issue_type, identifier_value, details) VALUES (?, ?, ?)",
                        (issue_type, identifier, ",".join(units)),
                    )
            connection.commit()
        os.replace(temporary_database, database_path)
    finally:
        workbook.close()
        temporary_database.unlink(missing_ok=True)
    return {
        "records_imported": len(imported),
        "manual_records_imported": manual_count,
        "ambiguous_plates": ambiguities["plate"],
        "ambiguous_vins": ambiguities["vin"],
        "duplicate_units": ambiguities["unit"],
    }


def find_units_by_identifier(
    database_path: str | Path,
    *,
    unit: Any = None,
    vin: Any = None,
    plate: Any = None,
) -> list[str]:
    clauses = []
    parameters = []
    if unit is not None:
        clauses.append("normalized_unit = ?")
        parameters.append(normalize_unit(unit))
    if vin is not None:
        clauses.append("vin = ?")
        parameters.append(normalize_vin(vin))
    if plate is not None:
        clauses.append("plate = ?")
        parameters.append(normalize_plate(plate))
    if not clauses:
        return []

    query = "SELECT DISTINCT normalized_unit FROM units WHERE " + " OR ".join(clauses)
    with closing(sqlite3.connect(database_path)) as connection:
        rows = connection.execute(query, parameters).fetchall()
    return sorted((row[0] for row in rows if row[0]), key=lambda value: (len(value), value))


def find_asset_owner(database_path: str | Path, unit: Any) -> str | None:
    normalized_unit = normalize_unit(unit)
    if not normalized_unit:
        return None
    with closing(sqlite3.connect(database_path)) as connection:
        columns = {row[1] for row in connection.execute("PRAGMA table_info(units)")}
        if "asset_owner" not in columns:
            return None
        rows = connection.execute(
            "SELECT DISTINCT asset_owner FROM units WHERE normalized_unit = ? AND asset_owner <> ''",
            (normalized_unit,),
        ).fetchall()
    owners = {row[0] for row in rows if row[0]}
    return next(iter(owners)) if len(owners) == 1 else None


MANAGEMENT_SCHEMA = """
CREATE TABLE IF NOT EXISTS custom_fields (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL COLLATE NOCASE UNIQUE,
    field_key TEXT NOT NULL UNIQUE,
    field_type TEXT NOT NULL CHECK(field_type IN ('text', 'date', 'number', 'yes_no')),
    created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS unit_custom_values (
    unit_id INTEGER NOT NULL REFERENCES units(id) ON DELETE CASCADE,
    field_id INTEGER NOT NULL REFERENCES custom_fields(id) ON DELETE CASCADE,
    value TEXT NOT NULL DEFAULT '',
    PRIMARY KEY (unit_id, field_id)
);
CREATE TABLE IF NOT EXISTS database_audit (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp_utc TEXT NOT NULL,
    event TEXT NOT NULL,
    unit_id INTEGER,
    details_json TEXT NOT NULL
);
"""

EDITABLE_UNIT_FIELDS = (
    "display_unit",
    "unit_type",
    "year",
    "make",
    "model",
    "vehicle_type",
    "plate",
    "vin",
    "fuel_type",
    "next_dot",
    "dot_status",
    "asset_owner",
)

WORKBOOK_HEADERS = {
    "Unit #": "display_unit",
    "Unit Type": "unit_type",
    "Year": "year",
    "Make": "make",
    "Model": "model",
    "Type": "vehicle_type",
    "Tag": "plate",
    "Vin": "vin",
    "Fuel Type": "fuel_type",
    "Next DOT": "next_dot",
    "DOT Status": "dot_status",
    "Asset Owner": "asset_owner",
}


def _management_connection(database_path: str | Path) -> sqlite3.Connection:
    connection = sqlite3.connect(database_path, timeout=30)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    connection.executescript(MANAGEMENT_SCHEMA)
    return connection


def ensure_management_schema(database_path: str | Path) -> None:
    with closing(_management_connection(database_path)) as connection:
        connection.commit()


def _field_key(name: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    if not key:
        raise ValueError("Enter a field name containing letters or numbers.")
    return key


def _add_custom_field(connection: sqlite3.Connection, name: object, field_type: object) -> tuple[dict, bool]:
    clean_name = str(name or "").strip()
    clean_type = str(field_type or "").strip().lower()
    if not clean_name or len(clean_name) > 60:
        raise ValueError("Custom field names must contain 1 to 60 characters.")
    if clean_type not in {"text", "date", "number", "yes_no"}:
        raise ValueError("Custom field type must be text, date, number, or yes/no.")
    existing = connection.execute(
        "SELECT id, name, field_key, field_type FROM custom_fields WHERE name = ? COLLATE NOCASE",
        (clean_name,),
    ).fetchone()
    if existing:
        if existing["field_type"] != clean_type:
            raise ValueError(f"Custom field {clean_name!r} already exists with a different type.")
        return dict(existing), False
    key = _field_key(clean_name)
    if connection.execute("SELECT 1 FROM custom_fields WHERE field_key = ?", (key,)).fetchone():
        raise ValueError("That custom field conflicts with an existing field name.")
    cursor = connection.execute(
        "INSERT INTO custom_fields(name, field_key, field_type, created_at) VALUES (?, ?, ?, ?)",
        (clean_name, key, clean_type, datetime.now(timezone.utc).isoformat()),
    )
    return {
        "id": cursor.lastrowid,
        "name": clean_name,
        "field_key": key,
        "field_type": clean_type,
    }, True


def add_custom_field(database_path: str | Path, name: object, field_type: object = "text") -> dict:
    with closing(_management_connection(database_path)) as connection:
        connection.execute("BEGIN IMMEDIATE")
        field, _created = _add_custom_field(connection, name, field_type)
        connection.execute(
            "INSERT INTO database_audit(timestamp_utc, event, details_json) VALUES (?, ?, ?)",
            (datetime.now(timezone.utc).isoformat(), "custom_field_added", json.dumps(field, sort_keys=True)),
        )
        connection.commit()
    return field


def list_custom_fields(database_path: str | Path) -> list[dict]:
    with closing(_management_connection(database_path)) as connection:
        rows = connection.execute(
            "SELECT id, name, field_key, field_type FROM custom_fields ORDER BY name COLLATE NOCASE"
        ).fetchall()
    return [dict(row) for row in rows]


def _unit_record(connection: sqlite3.Connection, unit_id: int) -> dict | None:
    row = connection.execute("SELECT * FROM units WHERE id = ?", (unit_id,)).fetchone()
    if row is None:
        return None
    record = dict(row)
    values = connection.execute(
        "SELECT field_id, value FROM unit_custom_values WHERE unit_id = ?",
        (unit_id,),
    ).fetchall()
    record["custom_values"] = {item["field_id"]: item["value"] for item in values}
    return record


def get_unit_record(database_path: str | Path, unit_id: int) -> dict:
    with closing(_management_connection(database_path)) as connection:
        record = _unit_record(connection, unit_id)
    if record is None:
        raise ValueError("The selected asset no longer exists.")
    return record


def list_unit_records(database_path: str | Path, search: object = "") -> list[dict]:
    query = str(search or "").strip().upper()
    with closing(_management_connection(database_path)) as connection:
        rows = connection.execute("SELECT id FROM units").fetchall()
        records = [_unit_record(connection, row["id"]) for row in rows]
    visible = [record for record in records if record is not None]
    if query:
        visible = [
            record
            for record in visible
            if query in " ".join(
                str(record.get(name) or "").upper()
                for name in ("display_unit", "unit_type", "make", "model", "vehicle_type", "plate", "vin", "asset_owner")
            )
        ]
    return sorted(visible, key=lambda record: (len(record["normalized_unit"]), record["normalized_unit"], record["id"]))


def _normalize_custom_value(value: object, field_type: str, field_name: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) > 500:
        raise ValueError(f"{field_name} is too long.")
    if field_type == "date":
        try:
            date.fromisoformat(text)
        except ValueError as error:
            raise ValueError(f"{field_name} must use YYYY-MM-DD.") from error
    elif field_type == "number":
        try:
            float(text)
        except ValueError as error:
            raise ValueError(f"{field_name} must be a number.") from error
    elif field_type == "yes_no":
        values = {"yes": "Yes", "true": "Yes", "1": "Yes", "no": "No", "false": "No", "0": "No"}
        if text.lower() not in values:
            raise ValueError(f"{field_name} must be Yes or No.")
        text = values[text.lower()]
    return text


def _validated_unit_values(
    connection: sqlite3.Connection,
    unit_id: int,
    current: dict,
    values: dict,
) -> dict:
    cleaned = {name: _text(values.get(name, current.get(name))) for name in EDITABLE_UNIT_FIELDS}
    if not re.fullmatch(r"\d{1,12}", cleaned["display_unit"]):
        raise ValueError("Unit number must contain 1 to 12 digits.")
    cleaned["normalized_unit"] = normalize_unit(cleaned["display_unit"])
    cleaned["plate"] = normalize_plate(cleaned["plate"])
    cleaned["vin"] = normalize_vin(cleaned["vin"])
    if cleaned["year"] and not re.fullmatch(r"\d{4}", cleaned["year"]):
        raise ValueError("Year must contain four digits.")
    if cleaned["asset_owner"] not in {"", "Little B's Asset", "Farm Asset"}:
        raise ValueError("Asset owner must be Little B's Asset, Farm Asset, or blank.")
    checks = (
        ("normalized_unit", "unit number is already assigned"),
        ("plate", "plate/tag is already assigned"),
        ("vin", "VIN/serial number is already assigned"),
    )
    for field, message in checks:
        if cleaned[field] and cleaned[field] != str(current.get(field) or ""):
            conflict = connection.execute(
                f"SELECT id FROM units WHERE {field} = ? AND id <> ? LIMIT 1",
                (cleaned[field], unit_id),
            ).fetchone()
            if conflict:
                raise ValueError(f"That {message} to another asset.")
    return cleaned


def _update_unit(
    connection: sqlite3.Connection,
    unit_id: int,
    values: dict,
    custom_values: dict[int, object],
) -> dict:
    current = _unit_record(connection, unit_id)
    if current is None:
        raise ValueError("The selected asset no longer exists.")
    cleaned = _validated_unit_values(connection, unit_id, current, values)
    assignments = ", ".join(f"{name} = ?" for name in (*EDITABLE_UNIT_FIELDS, "normalized_unit"))
    connection.execute(
        f"UPDATE units SET {assignments} WHERE id = ?",
        [*(cleaned[name] for name in EDITABLE_UNIT_FIELDS), cleaned["normalized_unit"], unit_id],
    )
    definitions = {
        row["id"]: row
        for row in connection.execute("SELECT id, name, field_type FROM custom_fields").fetchall()
    }
    for raw_field_id, raw_value in custom_values.items():
        field_id = int(raw_field_id)
        definition = definitions.get(field_id)
        if definition is None:
            raise ValueError("A custom field no longer exists.")
        value = _normalize_custom_value(raw_value, definition["field_type"], definition["name"])
        connection.execute(
            """INSERT INTO unit_custom_values(unit_id, field_id, value) VALUES (?, ?, ?)
               ON CONFLICT(unit_id, field_id) DO UPDATE SET value = excluded.value""",
            (unit_id, field_id, value),
        )
    return _unit_record(connection, unit_id)


def update_unit_record(
    database_path: str | Path,
    unit_id: int,
    values: dict,
    custom_values: dict[int, object] | None = None,
) -> dict:
    with closing(_management_connection(database_path)) as connection:
        connection.execute("BEGIN IMMEDIATE")
        updated = _update_unit(connection, unit_id, values, custom_values or {})
        connection.execute(
            "INSERT INTO database_audit(timestamp_utc, event, unit_id, details_json) VALUES (?, ?, ?, ?)",
            (datetime.now(timezone.utc).isoformat(), "unit_updated", unit_id, json.dumps(updated, sort_keys=True)),
        )
        connection.commit()
    return updated


def merge_fleet_workbook(workbook_path: str | Path, database_path: str | Path) -> dict[str, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [_text(cell.value) for cell in sheet[1]]
    if "Unit #" not in headers:
        workbook.close()
        raise ValueError("The import workbook must contain a Unit # column.")
    parsed: list[tuple[int, dict, dict[str, object]]] = []
    seen_units: set[str] = set()
    try:
        for source_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            source = dict(zip(headers, row))
            unit = normalize_unit(source.get("Unit #"))
            if not unit:
                if any(_text(value) for value in row):
                    raise ValueError(f"Row {source_row} has no valid Unit #.")
                continue
            if unit in seen_units:
                raise ValueError(f"Unit {unit} appears more than once in the import workbook.")
            seen_units.add(unit)
            builtins = {
                field: _text(source.get(header))
                for header, field in WORKBOOK_HEADERS.items()
                if header in headers
            }
            custom = {
                header.removeprefix("Custom:").strip(): source.get(header)
                for header in headers
                if header.startswith("Custom:") and header.removeprefix("Custom:").strip()
            }
            parsed.append((source_row, builtins, custom))
    finally:
        workbook.close()

    inserted = updated = custom_fields_added = 0
    with closing(_management_connection(database_path)) as connection:
        connection.execute("BEGIN IMMEDIATE")
        custom_definitions: dict[str, dict] = {}
        for _row, _builtins, custom in parsed:
            for name in custom:
                if name.casefold() not in custom_definitions:
                    definition, created = _add_custom_field(connection, name, "text")
                    custom_definitions[name.casefold()] = definition
                    custom_fields_added += int(created)
        for source_row, builtins, custom in parsed:
            normalized = normalize_unit(builtins["display_unit"])
            matches = connection.execute(
                "SELECT id FROM units WHERE normalized_unit = ?",
                (normalized,),
            ).fetchall()
            if len(matches) > 1:
                raise ValueError(f"Unit {normalized} has duplicate database records; import was cancelled.")
            if matches:
                unit_id = matches[0]["id"]
                current = _unit_record(connection, unit_id)
                changes = dict(builtins)
                changes["display_unit"] = current["display_unit"]
                _update_unit(
                    connection,
                    unit_id,
                    changes,
                    {custom_definitions[name.casefold()]["id"]: value for name, value in custom.items()},
                )
                updated += 1
            else:
                values = {name: "" for name in EDITABLE_UNIT_FIELDS}
                values.update(builtins)
                values["normalized_unit"] = normalized
                cursor = connection.execute(
                    """INSERT INTO units (
                        source_row, display_unit, normalized_unit, unit_type, year, make,
                        model, vehicle_type, plate, vin, fuel_type, next_dot, dot_status,
                        asset_owner, asset_source
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'import')""",
                    (source_row, values["display_unit"], normalized, values["unit_type"], values["year"],
                     values["make"], values["model"], values["vehicle_type"], normalize_plate(values["plate"]),
                     normalize_vin(values["vin"]), values["fuel_type"], values["next_dot"], values["dot_status"],
                     values["asset_owner"]),
                )
                _update_unit(
                    connection,
                    cursor.lastrowid,
                    values,
                    {custom_definitions[name.casefold()]["id"]: value for name, value in custom.items()},
                )
                inserted += 1
        summary = {"inserted": inserted, "updated": updated, "custom_fields_added": custom_fields_added}
        connection.execute(
            "INSERT INTO database_audit(timestamp_utc, event, details_json) VALUES (?, ?, ?)",
            (datetime.now(timezone.utc).isoformat(), "workbook_merged", json.dumps(summary, sort_keys=True)),
        )
        connection.commit()
    return summary


def export_fleet_workbook(database_path: str | Path, workbook_path: str | Path) -> Path:
    fields = list_custom_fields(database_path)
    records = list_unit_records(database_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fleet Assets"
    sheet.append([*WORKBOOK_HEADERS.keys(), *(f"Custom: {field['name']}" for field in fields)])
    for record in records:
        sheet.append([
            *(record.get(field, "") for field in WORKBOOK_HEADERS.values()),
            *(record["custom_values"].get(field["id"], "") for field in fields),
        ])
    destination = Path(workbook_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f".{destination.name}.{os.getpid()}.tmp")
    try:
        workbook.save(temporary)
        os.replace(temporary, destination)
    finally:
        workbook.close()
        temporary.unlink(missing_ok=True)
    return destination
