import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from dotdocs.tools_database import (
    add_tool_certification,
    create_tool,
    ensure_tools_schema,
    get_tool,
    list_tool_certifications,
    list_tools,
    export_tools_workbook,
    import_tools_workbook,
    update_tool,
)


def _tool(**overrides):
    values = {
        "tool_id": "CAL-001",
        "description": "Digital pressure gauge",
        "category": "Pressure",
        "manufacturer": "Fluke",
        "model": "700G",
        "serial_number": "SN-1001",
        "location": "Main Shop",
        "custodian": "Safety",
        "calibration_required": True,
        "calibration_interval_months": 12,
        "active": True,
        "notes": "Reference gauge",
    }
    values.update(overrides)
    return values


def test_creates_and_lists_normalized_tool_records(tmp_path):
    database = tmp_path / "fleet.db"
    ensure_tools_schema(database)

    created = create_tool(database, _tool())

    assert created["display_tool_id"] == "CAL-001"
    assert created["normalized_tool_id"] == "CAL001"
    assert created["normalized_serial_number"] == "SN1001"
    assert created["calibration_required"] is True
    assert list_tools(database) == [created]


def test_rejects_duplicate_tool_id_and_serial_without_partial_write(tmp_path):
    database = tmp_path / "fleet.db"
    create_tool(database, _tool())

    with pytest.raises(ValueError, match="Tool ID"):
        create_tool(database, _tool(tool_id="cal 001", serial_number="SN-2000"))
    with pytest.raises(ValueError, match="serial number"):
        create_tool(database, _tool(tool_id="CAL-002", serial_number="sn 1001"))

    assert len(list_tools(database)) == 1


def test_validates_tool_fields_and_positive_calibration_interval(tmp_path):
    database = tmp_path / "fleet.db"

    with pytest.raises(ValueError, match="description"):
        create_tool(database, _tool(description=""))
    with pytest.raises(ValueError, match="interval"):
        create_tool(database, _tool(calibration_interval_months=0))
    with pytest.raises(ValueError, match="Tool ID"):
        create_tool(database, _tool(tool_id="***"))


def test_updates_tool_and_deactivation_preserves_certification_history(tmp_path):
    database = tmp_path / "fleet.db"
    tool = create_tool(database, _tool())
    certificate = add_tool_certification(
        database,
        tool["id"],
        {
            "certificate_type": "Calibration",
            "certificate_number": "CERT-88",
            "performed_date": "2026-08-01",
            "due_date": "2027-08-01",
            "provider": "Tinker & Rasor",
            "result": "pass",
            "document_path": str(tmp_path / "certificate.pdf"),
            "document_sha256": "a" * 64,
            "notes": "Passed",
        },
    )

    updated = update_tool(database, tool["id"], {**tool, "active": False, "location": "Quarantine"})

    assert updated["active"] is False
    assert updated["location"] == "Quarantine"
    assert list_tool_certifications(database, tool["id"]) == [certificate]


def test_certification_history_supports_multiple_records_and_validates_dates(tmp_path):
    database = tmp_path / "fleet.db"
    tool = create_tool(database, _tool())
    first = add_tool_certification(
        database,
        tool["id"],
        {"certificate_type": "Calibration", "performed_date": "2025-08-01", "due_date": "2026-08-01", "result": "pass"},
    )
    second = add_tool_certification(
        database,
        tool["id"],
        {"certificate_type": "Certification", "performed_date": "2026-08-02", "due_date": "2027-08-02", "result": "fail"},
    )

    assert list_tool_certifications(database, tool["id"]) == [second, first]
    assert get_tool(database, tool["id"])["id"] == tool["id"]
    with pytest.raises(ValueError, match="before"):
        add_tool_certification(
            database,
            tool["id"],
            {"certificate_type": "Calibration", "performed_date": "2026-08-02", "due_date": "2026-08-01"},
        )


def test_tool_database_connections_release_file_for_immediate_rename(tmp_path):
    database = tmp_path / "fleet.db"
    create_tool(database, _tool())
    assert list_tools(database)

    renamed = tmp_path / "renamed.db"
    Path(database).rename(renamed)

    assert renamed.is_file()
    with sqlite3.connect(renamed) as connection:
        assert connection.execute("SELECT COUNT(*) FROM tools").fetchone()[0] == 1


def test_imports_tools_workbook_and_reports_duplicate_identifiers(tmp_path):
    source = tmp_path / "tools.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tools"
    sheet.append([
        "Tool ID", "Description", "Category", "Manufacturer", "Model", "Serial Number",
        "Location", "Custodian", "Calibration Required", "Interval Months", "Active", "Notes",
    ])
    sheet.append(["CAL-001", "Pressure gauge", "Pressure", "Fluke", "700G", "SN-1", "Shop", "Safety", "Yes", 12, "Yes", ""])
    sheet.append(["CAL-002", "Torque wrench", "Torque", "Proto", "J6014", "SN-2", "Shop", "Fleet", "Yes", 6, "Yes", ""])
    workbook.save(source)

    database = tmp_path / "fleet.db"
    result = import_tools_workbook(source, database)

    assert result == {"inserted": 2, "updated": 0}
    assert [tool["display_tool_id"] for tool in list_tools(database)] == ["CAL-001", "CAL-002"]

    duplicate = tmp_path / "duplicate.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Tool ID", "Description", "Serial Number"])
    sheet.append(["CAL-003", "First", "SAME"])
    sheet.append(["CAL-004", "Second", "same"])
    workbook.save(duplicate)
    with pytest.raises(ValueError, match="serial number"):
        import_tools_workbook(duplicate, database)
    assert len(list_tools(database)) == 2


def test_export_includes_tools_and_certification_history_worksheets(tmp_path):
    database = tmp_path / "fleet.db"
    tool = create_tool(database, _tool())
    add_tool_certification(
        database,
        tool["id"],
        {"certificate_type": "Calibration", "performed_date": "2026-08-01", "due_date": "2027-08-01", "result": "pass"},
    )

    exported = tmp_path / "tools-backup.xlsx"
    export_tools_workbook(database, exported)

    workbook = load_workbook(exported, read_only=True, data_only=True)
    assert workbook.sheetnames == ["Tools", "Certification History"]
    tool_rows = list(workbook["Tools"].iter_rows(values_only=True))
    history_rows = list(workbook["Certification History"].iter_rows(values_only=True))
    workbook.close()
    assert tool_rows[1][0] == "CAL-001"
    assert history_rows[1][0] == "CAL-001"
    assert history_rows[1][1] == "Calibration"
