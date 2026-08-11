import sqlite3

import pytest
from openpyxl import Workbook, load_workbook

from dotdocs.database import (
    add_custom_field,
    export_fleet_workbook,
    get_unit_record,
    import_fleet_workbook,
    list_custom_fields,
    list_unit_records,
    merge_fleet_workbook,
    update_unit_record,
)


def _database(tmp_path):
    source = tmp_path / "initial.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Unit #", "Unit Type", "Year", "Make", "Model", "Type", "Tag", "Vin", "Fuel Type", "Next DOT", "DOT Status", "Asset Owner"])
    sheet.append(["091", "Truck", "2018", "Ford", "F-350", "Truck", "OLD91", "1FT8W3BT0JEC00091", "Diesel", "2026-10-01", "Current", "Little B's Asset"])
    workbook.save(source)
    database = tmp_path / "fleet.db"
    import_fleet_workbook(source, database)
    return database


def test_custom_fields_and_asset_updates_are_persisted(tmp_path):
    database = _database(tmp_path)
    field = add_custom_field(database, "Registration Expires", "date")
    record = list_unit_records(database)[0]

    updated = update_unit_record(
        database,
        record["id"],
        {**record, "display_unit": "091", "make": "FORD", "plate": "NEW-91"},
        {field["id"]: "2027-08-31"},
    )

    assert updated["normalized_unit"] == "91"
    assert updated["plate"] == "NEW91"
    assert updated["custom_values"][field["id"]] == "2027-08-31"
    assert list_custom_fields(database) == [field]
    assert get_unit_record(database, record["id"]) == updated


def test_update_rejects_identifier_conflicts_without_partial_changes(tmp_path):
    database = _database(tmp_path)
    with sqlite3.connect(database) as connection:
        connection.execute(
            """INSERT INTO units (
                source_row, display_unit, normalized_unit, unit_type, year, make,
                model, vehicle_type, plate, vin, fuel_type, next_dot, dot_status,
                asset_owner, asset_source
            ) VALUES (0, '92', '92', '', '', '', '', '', 'PLATE92', 'VIN00092', '', '', '', '', 'manual')"""
        )
    original = list_unit_records(database)[0]

    with pytest.raises(ValueError, match="plate/tag is already assigned"):
        update_unit_record(database, original["id"], {**original, "plate": "plate 92"}, {})

    assert get_unit_record(database, original["id"])["plate"] == "OLD91"


def test_merge_import_updates_adds_custom_fields_and_exports_complete_backup(tmp_path):
    database = _database(tmp_path)
    incoming = tmp_path / "merge.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Unit #", "Unit Type", "Year", "Make", "Model", "Type", "Tag", "Vin", "Custom: Assigned Driver"])
    sheet.append(["91", "Truck", "2018", "Ford Updated", "F-350", "Truck", "OLD91", "1FT8W3BT0JEC00091", "Alex"])
    sheet.append(["92", "Trailer", "2024", "Big Tex", "14LP", "Trailer", "TAG92", "SERIAL92", "Sam"])
    workbook.save(incoming)

    result = merge_fleet_workbook(incoming, database)

    assert result == {"inserted": 1, "updated": 1, "custom_fields_added": 1}
    records = {item["normalized_unit"]: item for item in list_unit_records(database)}
    assigned_driver = list_custom_fields(database)[0]
    assert records["91"]["make"] == "Ford Updated"
    assert records["91"]["custom_values"][assigned_driver["id"]] == "Alex"
    assert records["92"]["custom_values"][assigned_driver["id"]] == "Sam"

    exported = tmp_path / "backup.xlsx"
    export_fleet_workbook(database, exported)
    backup = load_workbook(exported, read_only=True, data_only=True)
    rows = list(backup.active.iter_rows(values_only=True))
    backup.close()
    assert "Custom: Assigned Driver" in rows[0]
    assert {str(row[0]) for row in rows[1:]} == {"091", "92"}
